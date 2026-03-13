from django.db.models import Count, Avg, Q
from django.utils import timezone
import datetime


class DashboardAnalyticsService:

    @staticmethod
    def get_overview_stats():
        stats = {
            'total_students': 0,
            'total_teachers': 0,
            'total_courses': 0,
            'total_enrollments': 0,
        }
        try:
            from apps.students.models import Student
            from apps.teachers.models import Teacher
            from apps.courses.models import Course, Enrollment

            stats = {
                'total_students': Student.objects.filter(status='active').count(),
                'total_teachers': Teacher.objects.count(),
                'total_courses': Course.objects.filter(status='active').count(),
                'total_enrollments': Enrollment.objects.filter(status='active').count(),
            }
        except Exception:
            pass
        return stats

    @staticmethod
    def get_attendance_trend(days=30):
        result = {'labels': [], 'data': []}
        try:
            from apps.attendance.models import Attendance

            end_date = timezone.now().date()
            start_date = end_date - datetime.timedelta(days=days)

            for i in range(days):
                current_date = start_date + datetime.timedelta(days=i)
                total = Attendance.objects.filter(date=current_date).count()
                present = Attendance.objects.filter(date=current_date, status='present').count()

                percentage = round((present / total * 100), 1) if total > 0 else 0
                result['labels'].append(current_date.strftime('%d-%b'))
                result['data'].append(percentage)
        except Exception:
            pass
        return result

    @staticmethod
    def get_grade_distribution():
        distribution = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
        try:
            from apps.assignments.models import Submission

            submissions = Submission.objects.filter(
                status='graded',
                score__isnull=False
            ).values('score', 'assignment__max_score')

            for sub in submissions:
                if sub['assignment__max_score']:
                    pct = (sub['score'] / sub['assignment__max_score']) * 100
                    if pct >= 90:
                        distribution['A'] += 1
                    elif pct >= 80:
                        distribution['B'] += 1
                    elif pct >= 70:
                        distribution['C'] += 1
                    elif pct >= 60:
                        distribution['D'] += 1
                    else:
                        distribution['F'] += 1
        except Exception:
            pass
        return distribution

    @staticmethod
    def get_top_students(limit=10):
        result = []
        try:
            from apps.students.models import Student
            result = list(
                Student.objects
                .filter(status='active', gpa__isnull=False)
                .select_related('user')
                .order_by('-gpa')[:limit]
                .values('student_id', 'user__first_name', 'user__last_name', 'gpa', 'direction')
            )
        except Exception:
            pass
        return result

    @staticmethod
    def get_course_enrollment_stats():
        result = []
        try:
            from apps.courses.models import Course

            courses = Course.objects.filter(status='active').annotate(
                enrolled_count=Count('enrollment', filter=Q(enrollment__status='active'))
            ).select_related('teacher__user')

            for course in courses:
                fill_pct = round(
                    (course.enrolled_count / course.max_students * 100), 1
                ) if course.max_students > 0 else 0

                result.append({
                    'code': course.code,
                    'title': course.title,
                    'teacher': course.teacher.user.get_full_name() if course.teacher else '—',
                    'enrolled': course.enrolled_count,
                    'max': course.max_students,
                    'fill_pct': fill_pct,
                })
        except Exception:
            pass
        return result

    @staticmethod
    def get_low_attendance_students(threshold=75):
        result = []
        try:
            from apps.attendance.models import Attendance
            from apps.students.models import Student

            students = Student.objects.filter(status='active').select_related('user')

            for student in students:
                records = Attendance.objects.filter(student=student)
                total = records.count()
                if total == 0:
                    continue

                present = records.filter(status='present').count()
                late = records.filter(status='late').count()
                pct = (present + late * 0.5) / total * 100

                if pct < threshold:
                    result.append({
                        'student': student,
                        'attendance_pct': round(pct, 1),
                        'total_classes': total,
                        'missed': records.filter(status='absent').count(),
                    })

            result.sort(key=lambda x: x['attendance_pct'])
        except Exception:
            pass
        return result


class ExportService:

    @staticmethod
    def export_students_to_excel(queryset=None)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO

            if queryset is None:
                from apps.students.models import Student
                queryset = Student.objects.filter(status='active').select_related('user')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "O'quvchilar"

            headers = ['ID', 'Ism', 'Familiya', 'Email', 'Guruh', "Yo'nalish", "Kurs yili", 'GPA', 'Holat']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(fill_type='solid', fgColor="2563EB")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for row, student in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=student.student_id)
                ws.cell(row=row, column=2, value=student.user.first_name)
                ws.cell(row=row, column=3, value=student.user.last_name)
                ws.cell(row=row, column=4, value=student.user.email)
                ws.cell(row=row, column=5, value=student.group)
                ws.cell(row=row, column=6, value=student.direction)
                ws.cell(row=row, column=7, value=student.year_of_study)
                ws.cell(row=row, column=8, value=float(student.gpa) if student.gpa else 0)
                ws.cell(row=row, column=9, value=student.status)

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except ImportError:
            return None

    @staticmethod
    def export_attendance_to_csv(course_id, start_date=None, end_date=None):
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["O'quvchi ID", 'Ism', 'Sana', 'Holat', 'Izoh'])

        try:
            from apps.attendance.models import Attendance

            qs = Attendance.objects.filter(course_id=course_id).select_related('student__user')
            if start_date:
                qs = qs.filter(date__gte=start_date)
            if end_date:
                qs = qs.filter(date__lte=end_date)

            for record in qs.order_by('date', 'student__student_id'):
                writer.writerow([
                    record.student.student_id,
                    record.student.user.get_full_name(),
                    record.date.strftime('%Y-%m-%d'),
                    record.get_status_display(),
                    record.note,
                ])
        except Exception:
            pass

        return output.getvalue()